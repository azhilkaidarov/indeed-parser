# Модуль для сохранения данных в файл или базу данных
import pandas as pd
from job_data import job_listings
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def export_to_excel():
    if not job_listings:
        raise ValueError("job_listing is empty!")
    
    jobs_title = []
    companies_name = []
    jobs_locations = []
    jobs_work_type = []
    jobs_schedule = []
    jobs_benefits = []
    jobs_description = []
    jobs_salary = []

    for job in job_listings:
        jobs_title.append(job.title)
        companies_name.append(job.company_name)
        jobs_locations.append(job.location)
        jobs_work_type.append(job.work_type)
        jobs_schedule.append(job.schedule)
        jobs_benefits.append(job.benefits)
        jobs_description.append(job.description)
        jobs_salary.append(job.salary)

    data = {
        "Title" : jobs_title,
        "Company name" : companies_name,
        "Location" : jobs_locations,
        "Schedule" : jobs_schedule,
        "Work type" : jobs_work_type,
        "Benefits" : jobs_benefits,
        "Salary" : jobs_salary, 
        "Description" : jobs_description
    }
    
    df = pd.DataFrame(data)
    
    # Сохранение в Excel с использованием openpyxl как движка
    with pd.ExcelWriter('test.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    
    # Загружаем рабочую книгу и активный лист
    workbook = load_workbook('test.xlsx')
    worksheet = workbook.active

    # Устанавливаем перенос текста для всех ячеек и автоматически изменяем высоту строки
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Определяем максимальную высоту строки, основанную на содержимом ячеек
        max_height = max([len(str(cell.value).split('\n')) for cell in row])
        worksheet.row_dimensions[row[0].row].height = max_height * 15  # Умножаем на 15 для лучшего соответствия высоты строки

    # Автоматически устанавливаем ширину столбцов
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Получаем буквенное обозначение столбца

        if column == "H": 
            worksheet.column_dimensions[column].width = 250
            continue

        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    # Сохраняем изменения
    workbook.save('test.xlsx')
    workbook.close()
